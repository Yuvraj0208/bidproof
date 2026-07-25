"""US-05 integration: the matrix export is a real xlsx with every row
carrying verdict + proof + confidence, and needs-human rows marked."""

import io

import pytest
from openpyxl import load_workbook

from test_checking_api import client_for, make_app, seed_and_upload
from test_parser_ladder import DIGITAL
from test_rules_api import FakeGateway
from test_upload_api import create_org

pytestmark = pytest.mark.integration


async def test_matrix_export_is_valid_xlsx_with_complete_rows(owner_conn):
    org_id = await create_org(owner_conn)
    app = make_app(FakeGateway(['{"rules": []}']))

    async with client_for(app, org_id) as client:
        tender_id = await seed_and_upload(client)
        await client.post(f"/tenders/{tender_id}/extract")
        await client.post(f"/tenders/{tender_id}/check")
        verdicts = (await client.get(f"/tenders/{tender_id}/verdicts")).json()
        export = await client.get(f"/tenders/{tender_id}/matrix.xlsx")

    assert export.status_code == 200
    assert export.headers["content-type"].endswith("spreadsheetml.sheet")

    workbook = load_workbook(io.BytesIO(export.content))
    sheet = workbook["Compliance Matrix"]
    headers = [cell.value for cell in sheet[1]]
    assert headers[:6] == ["Family", "Key", "Requirement", "Value", "Verdict", "Status"]
    assert sheet.max_row == len(verdicts) + 1  # header + one row per verdict

    for row in sheet.iter_rows(min_row=2, values_only=True):
        family, key, requirement, _value, verdict, status, reason, conf, band, *_rest, page, el_id = row
        assert family and key and requirement and verdict and reason
        assert 0 <= conf <= 1 and band in ("green", "yellow", "red")
        assert page >= 1 and el_id  # every row clicks back to a real element


async def test_matrix_marks_needs_human_rows_queued(owner_conn):
    org_id = await create_org(owner_conn)  # no capability data → needs_human rows
    app = make_app(FakeGateway(['{"rules": []}']))

    async with client_for(app, org_id) as client:
        response = await client.post(
            "/tenders/upload",
            files={"file": ("tender.pdf", DIGITAL, "application/pdf")},
        )
        tender_id = response.json()["tender_id"]
        await client.post(f"/tenders/{tender_id}/extract")
        await client.post(f"/tenders/{tender_id}/check")
        export = await client.get(f"/tenders/{tender_id}/matrix.xlsx")

    sheet = load_workbook(io.BytesIO(export.content))["Compliance Matrix"]
    statuses = [row[5] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert "QUEUED FOR HUMAN" in statuses


async def test_matrix_export_404_before_check(owner_conn):
    org_id = await create_org(owner_conn)
    app = make_app(FakeGateway([]))
    async with client_for(app, org_id) as client:
        import uuid

        response = await client.get(f"/tenders/{uuid.uuid4()}/matrix.xlsx")
    assert response.status_code == 404
