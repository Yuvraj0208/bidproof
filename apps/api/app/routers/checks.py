"""Checking endpoints (SPEC §5.5): run the Matcher + RiskScorer, read
verdicts (joined to their rule + element for click-to-proof) and risks,
and export the Compliance Matrix to Excel (US-05 — the money table)."""

import io
import uuid

from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Response
from pydantic import BaseModel, Field
from sqlalchemy import select

from app.core.db import org_scoped_session
from app.core.roles import Role, require_role
from app.core.tenancy import require_org_id
from app.models import AuditLog, Element, RiskRow, Rule, VerdictRow
from app.services import checking

router = APIRouter()


class BBoxOut(BaseModel):
    x0: float
    y0: float
    x1: float
    y1: float


class VerdictOut(BaseModel):
    id: uuid.UUID
    rule_id: uuid.UUID
    family: str
    key: str
    requirement_text: str
    value: str | None
    verdict: str
    reason: str
    confidence: float
    band: str
    arithmetic: bool
    cited_fact_id: uuid.UUID | None
    cited_product_id: uuid.UUID | None
    el_id: uuid.UUID
    document_id: uuid.UUID
    page_no: int
    bbox: BBoxOut
    # Set once a human has settled this verdict (SPEC §7 checkpoint 3).
    system_verdict: str | None = None
    decided_by: str | None = None
    decided_at: datetime | None = None
    decided_reason: str | None = None


class RiskOut(BaseModel):
    id: uuid.UUID
    code: str
    severity: str
    message: str
    rupee_impact: float | None
    el_id: uuid.UUID | None


@router.post("/tenders/{tender_id}/check")
async def run_check(
    tender_id: uuid.UUID,
    org_id: uuid.UUID = Depends(require_org_id),
    gateway=Depends(checking.get_checking_gateway),
) -> dict:
    summary = await checking.run_checks(org_id, tender_id, gateway=gateway)
    if summary is None:
        raise HTTPException(404, "tender not found")
    return summary


@router.get("/tenders/{tender_id}/verdicts", response_model=list[VerdictOut])
async def list_verdicts(
    tender_id: uuid.UUID, org_id: uuid.UUID = Depends(require_org_id)
) -> list[VerdictOut]:
    async with org_scoped_session(org_id) as session:
        rows = (
            await session.execute(
                select(VerdictRow, Rule, Element)
                .join(Rule, VerdictRow.rule_id == Rule.rule_id)
                .join(Element, Rule.el_id == Element.el_id)
                .where(VerdictRow.tender_id == tender_id)
                .order_by(Rule.family, Rule.key)
            )
        ).all()
        return [
            VerdictOut(
                id=v.id, rule_id=v.rule_id, family=r.family, key=r.key,
                requirement_text=r.requirement_text, value=r.value_text,
                verdict=v.verdict, reason=v.reason, confidence=v.confidence,
                band=v.band, arithmetic=v.arithmetic,
                cited_fact_id=v.cited_fact_id, cited_product_id=v.cited_product_id,
                el_id=e.el_id, document_id=e.document_id, page_no=e.page_no,
                bbox=BBoxOut(x0=e.x0, y0=e.y0, x1=e.x1, y1=e.y1),
                system_verdict=v.system_verdict, decided_by=v.decided_by,
                decided_at=v.decided_at, decided_reason=v.decided_reason,
            )
            for v, r, e in rows
        ]


MATRIX_HEADERS = [
    "Family", "Key", "Requirement", "Value", "Verdict", "Status", "Reason",
    "Confidence", "Band", "Arithmetic", "Page", "Proof (el_id)",
]


@router.get("/tenders/{tender_id}/matrix.xlsx")
async def export_matrix(
    tender_id: uuid.UUID, org_id: uuid.UUID = Depends(require_org_id)
) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    async with org_scoped_session(org_id) as session:
        rows = (
            await session.execute(
                select(VerdictRow, Rule, Element)
                .join(Rule, VerdictRow.rule_id == Rule.rule_id)
                .join(Element, Rule.el_id == Element.el_id)
                .where(VerdictRow.tender_id == tender_id)
                .order_by(Rule.family, Rule.key)
            )
        ).all()
    if not rows:
        raise HTTPException(404, "no verdicts for this tender — run /check first")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Compliance Matrix"
    sheet.append(MATRIX_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    queued_fill = PatternFill("solid", start_color="FFF3CD")
    for verdict, rule, element in rows:
        status = "QUEUED FOR HUMAN" if verdict.verdict == "needs_human" else "decided"
        sheet.append([
            rule.family, rule.key, rule.requirement_text, rule.value_text,
            verdict.verdict, status, verdict.reason,
            round(verdict.confidence, 2), verdict.band,
            "yes" if verdict.arithmetic else "no",
            element.page_no, str(rule.el_id),
        ])
        if verdict.verdict == "needs_human":
            for cell in sheet[sheet.max_row]:
                cell.fill = queued_fill

    buffer = io.BytesIO()
    workbook.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="matrix-{tender_id}.xlsx"'
        },
    )


@router.get("/tenders/{tender_id}/risks", response_model=list[RiskOut])
async def list_risks(
    tender_id: uuid.UUID, org_id: uuid.UUID = Depends(require_org_id)
) -> list[RiskOut]:
    async with org_scoped_session(org_id) as session:
        rows = (
            await session.execute(
                select(RiskRow)
                .where(RiskRow.tender_id == tender_id)
                .order_by(RiskRow.severity.desc(), RiskRow.code)
            )
        ).scalars()
        return [
            RiskOut(
                id=r.id, code=r.code, severity=r.severity, message=r.message,
                rupee_impact=float(r.rupee_impact) if r.rupee_impact is not None else None,
                el_id=r.el_id,
            )
            for r in rows
        ]


# The vocabulary a human may choose from. `needs_human` is absent on purpose:
# this endpoint exists to LEAVE that state, and re-asserting it would be a way to
# make the review queue look clear while nothing was decided.
HUMAN_VERDICTS = ("complies", "partial", "gap", "not_applicable")


class DecideVerdictIn(BaseModel):
    verdict: str
    # A verdict a human asserted with no reason is not evidence. The compliance
    # matrix is the artefact a bid is defended with, so the reason is required.
    reason: str = Field(min_length=3, max_length=2000)
    name: str = Field(min_length=2, max_length=120)


@router.post(
    "/tenders/{tender_id}/verdicts/{verdict_id}/decide", response_model=VerdictOut
)
async def decide_verdict(
    tender_id: uuid.UUID,
    verdict_id: uuid.UUID,
    body: DecideVerdictIn,
    org_id: uuid.UUID = Depends(require_org_id),
    role: Role = Depends(require_role(Role.REVIEWER)),
) -> VerdictOut:
    """Settle a verdict the system refused to guess (SPEC §7 checkpoint 3).

    The matrix could always say `needs_human` — "no arithmetic and no cited
    judgement could settle this" — and the Review Hub counted those as blocking
    submission, but there was no way to answer. This is that way.

    The machine's verdict is preserved in `system_verdict`, so the override stays
    visible and a human answer is never mistaken for a grounded machine one. The
    rule's own proof (page + box) is untouched — the human is deciding what a
    cited requirement MEANS for this company, not inventing the requirement.
    """
    if body.verdict not in HUMAN_VERDICTS:
        raise HTTPException(
            400,
            f"verdict must be one of {list(HUMAN_VERDICTS)} — "
            "'needs_human' is what you are resolving, not a resolution",
        )

    async with org_scoped_session(org_id) as session:
        verdict = await session.get(VerdictRow, verdict_id)
        if verdict is None or verdict.tender_id != tender_id:
            raise HTTPException(404, "verdict not found")

        previous = verdict.verdict
        # Keep the FIRST machine verdict, not the previous human one, so a
        # second correction still shows what the system originally said.
        if verdict.system_verdict is None:
            verdict.system_verdict = previous
        verdict.verdict = body.verdict
        verdict.reason = f"decided by {body.name}: {body.reason}"
        verdict.decided_by = body.name
        verdict.decided_at = datetime.now(timezone.utc)
        verdict.decided_reason = body.reason
        # A human decision is certain by construction — they are the authority
        # here — but it is NOT a model confidence, which is why the band goes
        # green only through the explicit human route and `decided_by` is what
        # the UI keys on to say so.
        verdict.confidence = 1.0
        verdict.band = "green"
        rule_id = verdict.rule_id

        session.add(AuditLog(
            org_id=org_id, tender_id=tender_id, actor=body.name,
            action="verdict_decided_by_human",
            details={
                "verdict_id": str(verdict_id), "rule_id": str(rule_id),
                "from": previous, "to": body.verdict,
                "reason": body.reason, "role": role.value,
            },
        ))

    async with org_scoped_session(org_id) as session:
        row = (
            await session.execute(
                select(VerdictRow, Rule, Element)
                .join(Rule, VerdictRow.rule_id == Rule.rule_id)
                .join(Element, Rule.el_id == Element.el_id)
                .where(VerdictRow.id == verdict_id)
            )
        ).first()
        v, r, e = row
        return VerdictOut(
            id=v.id, rule_id=v.rule_id, family=r.family, key=r.key,
            requirement_text=r.requirement_text, value=r.value_text,
            verdict=v.verdict, reason=v.reason, confidence=v.confidence,
            band=v.band, arithmetic=v.arithmetic,
            cited_fact_id=v.cited_fact_id, cited_product_id=v.cited_product_id,
            el_id=e.el_id, document_id=e.document_id, page_no=e.page_no,
            bbox=BBoxOut(x0=e.x0, y0=e.y0, x1=e.x1, y1=e.y1),
            system_verdict=v.system_verdict, decided_by=v.decided_by,
            decided_at=v.decided_at, decided_reason=v.decided_reason,
        )
